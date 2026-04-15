"""Build XLSX artifacts — Operational KPI Dashboard and Financial Model Summary."""
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.header_footer import HeaderFooter

# Brand
NAVY = "FF002060"
ORANGE = "FFFC7134"
GREEN = "FF00B050"
LIGHT_NAVY = "FFE6EBF4"
WHITE = "FFFFFFFF"
BLACK = "FF000000"
GREY = "FF595959"
RED = "FFC00000"
YELLOW_FILL = "FFFFF2CC"
GREEN_FILL = "FFE2EFDA"
RED_FILL = "FFFCE4E4"

FONT_NAME = "Segoe UI"

MISSION_FOOTER = ("NextDecade Corporation (NextDecade) is committed to providing the world access "
                  "to lower carbon intensive energy through Rio Grande LNG & NEXT Carbon Solutions.")
CLASSIFICATION_FOOTER = ("Confidential and Proprietary - This document is intended solely for internal use. "
                        "Unauthorized disclosure, distribution, or reproduction is strictly prohibited. "
                        "Content is preliminary and subject to revision.")


def _hdr_style():
    return {
        "font": Font(name=FONT_NAME, size=11, bold=True, color=WHITE),
        "fill": PatternFill("solid", fgColor=NAVY),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _thin_border():
    s = Side(style="thin", color="FFBFBFBF")
    return Border(top=s, bottom=s, left=s, right=s)


def _apply_style(cell, style):
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]


def _body_font(bold=False, color=BLACK):
    return Font(name=FONT_NAME, size=10, bold=bold, color=color)


def _set_title_row(ws, row, text, n_cols, color=NAVY):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1)
    c.font = Font(name=FONT_NAME, size=16, bold=True, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _set_subtitle_row(ws, row, text, n_cols, color=ORANGE):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1)
    c.font = Font(name=FONT_NAME, size=10, italic=True, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _set_chrome(ws):
    """Apply NDLNG chrome: footer mission line + classification footer, page margins."""
    ws.page_margins = PageMargins(left=0.75, right=0.75, top=0.75, bottom=0.9,
                                  header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True
    ws.sheet_view.showGridLines = False
    hf: HeaderFooter = ws.HeaderFooter
    hf.oddHeader.center.text = "&\"Segoe UI\"&10&KFC7134" + "NextDecade Corporation"
    hf.oddFooter.left.text = "&\"Segoe UI\"&8&K595959" + MISSION_FOOTER
    hf.oddFooter.center.text = "&\"Segoe UI\"&8&K595959" + CLASSIFICATION_FOOTER
    hf.oddFooter.right.text = "&\"Segoe UI\"&8&K595959" + "Page &P of &N"


def build_kpi_dashboard(output: Path):
    wb = Workbook()
    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"

    _set_title_row(ws, 1, "Rio Grande LNG - Construction KPI Dashboard", 7)
    _set_subtitle_row(ws, 2, "Period ending March 31, 2026  |  Data as of April 15, 2026", 7)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # Summary KPI cards (row 4 headers, row 5 values)
    cards = [
        ("Construction % Complete", "68%", "Target 67% - ahead by 1 pt"),
        ("TRIR (Rolling 12 mo.)", "0.42", "Below industry benchmark 0.65"),
        ("Project Spend ($B)", "$9.8B", "Budget: $18.4B total - 53%"),
        ("Schedule Confidence", "High", "First LNG: 2H 2027"),
    ]
    for i, (label, value, note) in enumerate(cards):
        col = 1 + i * 2
        ws.cell(row=4, column=col, value=label)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        c = ws.cell(row=4, column=col)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=5, column=col, value=value)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        c = ws.cell(row=5, column=col)
        c.font = Font(name=FONT_NAME, size=22, bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=6, column=col, value=note)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        c = ws.cell(row=6, column=col)
        c.font = Font(name=FONT_NAME, size=9, italic=True, color=GREY)
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[5].height = 38
    ws.row_dimensions[6].height = 22

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    _set_chrome(ws)

    # --- Sheet 2: Progress by Area ---
    ws2 = wb.create_sheet("Progress by Area")
    _set_title_row(ws2, 1, "Rio Grande LNG - Progress by Area", 5)
    _set_subtitle_row(ws2, 2, "All values in percent complete", 5)

    headers = ["Area", "Engineering", "Procurement", "Construction", "Status"]
    for i, h in enumerate(headers, 1):
        c = ws2.cell(row=4, column=i, value=h)
        _apply_style(c, _hdr_style())
    ws2.row_dimensions[4].height = 22

    rows = [
        ("Train 1",  100, 96, 72, "On track"),
        ("Train 2",  100, 92, 55, "On track"),
        ("Train 3",   99, 82, 34, "On track"),
        ("Common Utilities", 100, 98, 78, "On track"),
        ("Storage Tanks",    100, 95, 62, "On track"),
        ("Marine / Jetty",    98, 72, 38, "Monitor"),
        ("Feed Gas Pipeline", 100, 99, 88, "On track"),
    ]
    for ridx, row in enumerate(rows, start=5):
        for cidx, v in enumerate(row, start=1):
            c = ws2.cell(row=ridx, column=cidx, value=v)
            c.font = _body_font(bold=(cidx == 1))
            c.alignment = Alignment(horizontal="center" if cidx > 1 else "left",
                                    vertical="center")
            c.border = _thin_border()
            if cidx == 5:  # status
                if v == "Monitor":
                    c.fill = PatternFill("solid", fgColor=YELLOW_FILL)
                elif v == "On track":
                    c.fill = PatternFill("solid", fgColor=GREEN_FILL)
                else:
                    c.fill = PatternFill("solid", fgColor=RED_FILL)

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # Chart: construction % by area
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Construction % Complete by Area"
    data = Reference(ws2, min_col=4, min_row=4, max_row=4 + len(rows))
    cats = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 2
    ws2.add_chart(chart, "G4")

    _set_chrome(ws2)

    # --- Sheet 3: HSSE Trend ---
    ws3 = wb.create_sheet("HSSE Trend")
    _set_title_row(ws3, 1, "HSSE - 12 Month Trend", 4)
    _set_subtitle_row(ws3, 2, "TRIR = Total Recordable Incident Rate (per 200,000 hours)", 4)

    hsse_hdr = ["Month", "Hours (M)", "Recordables", "TRIR"]
    for i, h in enumerate(hsse_hdr, 1):
        c = ws3.cell(row=4, column=i, value=h)
        _apply_style(c, _hdr_style())

    months = [
        ("Apr 2025", 2.4, 6, 0.50),
        ("May 2025", 2.5, 5, 0.40),
        ("Jun 2025", 2.6, 7, 0.54),
        ("Jul 2025", 2.8, 6, 0.43),
        ("Aug 2025", 2.9, 5, 0.34),
        ("Sep 2025", 3.0, 7, 0.47),
        ("Oct 2025", 3.0, 6, 0.40),
        ("Nov 2025", 2.9, 5, 0.34),
        ("Dec 2025", 2.7, 6, 0.44),
        ("Jan 2026", 2.9, 7, 0.48),
        ("Feb 2026", 3.0, 6, 0.40),
        ("Mar 2026", 3.1, 7, 0.45),
    ]
    for ridx, row in enumerate(months, start=5):
        for cidx, v in enumerate(row, start=1):
            c = ws3.cell(row=ridx, column=cidx, value=v)
            c.font = _body_font(bold=(cidx == 1))
            c.alignment = Alignment(horizontal="center" if cidx > 1 else "left",
                                    vertical="center")
            c.border = _thin_border()
            if cidx == 4:
                c.number_format = "0.00"

    # Totals row
    total_row = 5 + len(months)
    ws3.cell(row=total_row, column=1, value="TOTAL / Avg").font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
    ws3.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row-1})").font = _body_font(bold=True)
    ws3.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})").font = _body_font(bold=True)
    ws3.cell(row=total_row, column=4, value=f"=AVERAGE(D5:D{total_row-1})").font = _body_font(bold=True)
    for c in range(1, 5):
        ws3.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        ws3.cell(row=total_row, column=c).border = _thin_border()
    ws3.cell(row=total_row, column=4).number_format = "0.00"

    # Chart
    chart = LineChart()
    chart.title = "TRIR - 12 Month Rolling"
    data = Reference(ws3, min_col=4, min_row=4, max_row=4 + len(months))
    cats = Reference(ws3, min_col=1, min_row=5, max_row=4 + len(months))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 13
    ws3.add_chart(chart, "F4")

    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    _set_chrome(ws3)

    wb.save(str(output))
    print(f"Wrote {output}")


def build_financial_model(output: Path):
    wb = Workbook()

    # Assumptions sheet
    ws = wb.active
    ws.title = "Assumptions"
    _set_title_row(ws, 1, "Phase 2 Financial Model - Assumptions", 4)
    _set_subtitle_row(ws, 2, "Illustrative only - not an offering", 4)

    headers = ["Parameter", "Value", "Units", "Notes"]
    for i, h in enumerate(headers, 1):
        _apply_style(ws.cell(row=4, column=i, value=h), _hdr_style())

    rows = [
        ("Phase 2 Capex",               6.2, "$B", "Illustrative"),
        ("Train count (Phase 2)",         2, "#",  ""),
        ("Nameplate capacity per train",  6.0, "MTPA", ""),
        ("Opex per tonne",               35,   "$/t", "Annualized"),
        ("Henry Hub price (flat)",       3.75, "$/MMBtu", "Flat assumption"),
        ("SPA fixed fee",                 2.50, "$/MMBtu", ""),
        ("SPA variable (115% HH + fee)",  "=E7*1.15+E8", "$/MMBtu", "HH linked"),
        ("WACC",                          8.0, "%",     ""),
        ("Contract term",                  20, "years", ""),
        ("Tax rate",                       21, "%",     ""),
    ]
    for ridx, row in enumerate(rows, start=5):
        for cidx, v in enumerate(row, start=1):
            c = ws.cell(row=ridx, column=cidx, value=v)
            c.font = _body_font()
            c.alignment = Alignment(horizontal="left" if cidx in (1, 3, 4) else "right",
                                    vertical="center")
            c.border = _thin_border()
            if cidx == 2 and isinstance(v, (int, float)):
                c.number_format = "#,##0.00"

    for col, w in zip([1, 2, 3, 4], [36, 14, 10, 48]):
        ws.column_dimensions[get_column_letter(col)].width = w
    _set_chrome(ws)

    # Capex waterfall (simple)
    ws2 = wb.create_sheet("Capex")
    _set_title_row(ws2, 1, "Phase 2 Capex Buildup", 3)
    _set_subtitle_row(ws2, 2, "Illustrative", 3)

    for i, h in enumerate(["Category", "Capex ($B)", "% of Total"], 1):
        _apply_style(ws2.cell(row=4, column=i, value=h), _hdr_style())

    cat = [
        ("EPC - Liquefaction",            3.6),
        ("EPC - Storage & Marine",        0.8),
        ("Owner's Costs",                 0.4),
        ("Development & Permitting",      0.2),
        ("Contingency",                   0.7),
        ("Financing Costs (capitalized)", 0.5),
    ]
    total_capex = sum(v for _, v in cat)
    for ridx, (k, v) in enumerate(cat, start=5):
        ws2.cell(row=ridx, column=1, value=k).font = _body_font(bold=True)
        c = ws2.cell(row=ridx, column=2, value=v)
        c.font = _body_font()
        c.number_format = "$#,##0.0"
        c.alignment = Alignment(horizontal="right")
        c.border = _thin_border()
        c2 = ws2.cell(row=ridx, column=3, value=f"=B{ridx}/SUM($B$5:$B${4+len(cat)})")
        c2.number_format = "0.0%"
        c2.font = _body_font()
        c2.alignment = Alignment(horizontal="right")
        c2.border = _thin_border()
        ws2.cell(row=ridx, column=1).border = _thin_border()

    tr = 5 + len(cat)
    ws2.cell(row=tr, column=1, value="TOTAL").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    ws2.cell(row=tr, column=1).fill = PatternFill("solid", fgColor=NAVY)
    tot = ws2.cell(row=tr, column=2, value=f"=SUM(B5:B{tr-1})")
    tot.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    tot.fill = PatternFill("solid", fgColor=NAVY)
    tot.number_format = "$#,##0.0"
    tot.alignment = Alignment(horizontal="right")
    ws2.cell(row=tr, column=3, value=1).number_format = "0.0%"
    ws2.cell(row=tr, column=3).font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    ws2.cell(row=tr, column=3).fill = PatternFill("solid", fgColor=NAVY)
    ws2.cell(row=tr, column=3).alignment = Alignment(horizontal="right")

    # Chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Phase 2 Capex by Category ($B)"
    data = Reference(ws2, min_col=2, min_row=4, max_row=4 + len(cat))
    cats = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(cat))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 2
    ws2.add_chart(chart, "E4")

    for col, w in zip([1, 2, 3], [36, 14, 14]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    _set_chrome(ws2)

    # Operating metrics
    ws3 = wb.create_sheet("Ops Summary")
    _set_title_row(ws3, 1, "Phase 2 Annual Operating Summary (Illustrative)", 4)
    _set_subtitle_row(ws3, 2, "Per train, steady state", 4)

    for i, h in enumerate(["Metric", "Per Train", "Phase 2 Total", "Notes"], 1):
        _apply_style(ws3.cell(row=4, column=i, value=h), _hdr_style())

    ops = [
        ("Nameplate capacity (MTPA)",        6.0, 12.0, ""),
        ("Expected availability (%)",        95,  95,   "TMY"),
        ("Deliverable LNG (MTPA)",           5.7, 11.4, ""),
        ("Henry Hub feed gas (MMBtu, B)",    320, 640,  "Illustrative"),
        ("Revenue ($B, illustrative)",       2.8, 5.6,  "HH + fee"),
        ("Opex ($B)",                        0.20, 0.40, ""),
        ("EBITDA ($B)",                      2.60, 5.20, ""),
    ]
    for ridx, row in enumerate(ops, start=5):
        for cidx, v in enumerate(row, start=1):
            c = ws3.cell(row=ridx, column=cidx, value=v)
            c.font = _body_font(bold=(cidx == 1))
            c.alignment = Alignment(horizontal="left" if cidx in (1, 4) else "right",
                                    vertical="center")
            c.border = _thin_border()
            if cidx in (2, 3) and isinstance(v, (int, float)):
                c.number_format = "#,##0.00"

    for col, w in zip([1, 2, 3, 4], [34, 14, 16, 30]):
        ws3.column_dimensions[get_column_letter(col)].width = w
    _set_chrome(ws3)

    wb.save(str(output))
    print(f"Wrote {output}")


if __name__ == "__main__":
    out_dir = Path(sys.argv[1])
    out_dir.mkdir(parents=True, exist_ok=True)
    build_kpi_dashboard(out_dir / "Construction KPI Dashboard.xlsx")
    build_financial_model(out_dir / "Phase 2 Financial Model Summary.xlsx")
